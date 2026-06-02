from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views import View
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
import json

from .models import *
from .permissions import *


def health_check(request):
    return JsonResponse({'status': 'ok'})


# ========== VISTAS DE AUTENTICACIÓN ==========

class LoginView(View):
    def get(self, request):
        if request.user.is_authenticated:
            return redirect('dashboard')
        return render(request, 'login.html')
    
    def post(self, request):
        documento = request.POST.get('documento')
        password = request.POST.get('password')
        user = authenticate(request, documento=documento, password=password)
        
        if user is not None and user.activo:
            login(request, user)
            return redirect('dashboard')
        else:
            return render(request, 'login.html', {
                'error': 'Documento o contraseña incorrectos, o usuario inactivo'
            })


class LogoutView(View):
    def get(self, request):
        logout(request)
        return redirect('login')


@method_decorator(login_required, name='dispatch')
class CambiarContrasenaView(View):
    def get(self, request):
        return render(request, 'cambiar_contrasena.html')
    
    def post(self, request):
        contrasena_actual = request.POST.get('contrasena_actual')
        contrasena_nueva = request.POST.get('contrasena_nueva')
        contrasena_confirmar = request.POST.get('contrasena_confirmar')
        
        # Verificar contraseña actual
        user = authenticate(request, documento=request.user.documento, password=contrasena_actual)
        if user is None:
            return render(request, 'cambiar_contrasena.html', {
                'error': 'La contraseña actual es incorrecta'
            })
        
        # Verificar que las nuevas contraseñas coincidan
        if contrasena_nueva != contrasena_confirmar:
            return render(request, 'cambiar_contrasena.html', {
                'error': 'Las nuevas contraseñas no coinciden'
            })
        
        # Verificar longitud mínima
        if len(contrasena_nueva) < 6:
            return render(request, 'cambiar_contrasena.html', {
                'error': 'La contraseña debe tener al menos 6 caracteres'
            })
        
        # Cambiar contraseña
        request.user.set_password(contrasena_nueva)
        request.user.save()
        
        # Re-autenticar con la nueva contraseña
        login(request, request.user)
        
        return render(request, 'cambiar_contrasena.html', {
            'success': 'Contraseña cambiada exitosamente'
        })


@method_decorator(login_required, name='dispatch')
class DashboardView(View):
    def get(self, request):
        user = request.user
        
        if user.rol == 'admin_general':
            # Datos para admin general - Solo lista de empresarios
            empresarios = Usuario.objects.filter(rol='empresario').order_by('-created_at')
            
            context = {
                'ultimos_empresarios': empresarios,
            }
            return render(request, 'admin_dashboard.html', context)
        
        elif user.rol == 'empresario':
            # Datos para empresario
            ahora_local = timezone.localtime(timezone.now())
            hoy = ahora_local.date()
            
            # Chanceros
            total_chanceros = Usuario.objects.filter(
                rol='chancero',
                empresario=user
            ).count()
            chanceros_activos = Usuario.objects.filter(
                rol='chancero',
                empresario=user,
                activo=True
            ).count()
            
            # Ventas de hoy (usando fecha local)
            apuestas_hoy = Apuesta.objects.filter(
                empresario=user,
                fecha_hora__date=hoy
            )
            ventas_hoy = apuestas_hoy.aggregate(total=Sum('monto_apostado'))['total'] or 0
            apuestas_count = apuestas_hoy.count()
            
            # Premios pagados hoy
            premios_pagados_hoy = apuestas_hoy.filter(estado='pagada').aggregate(
                total=Sum('premio_potencial')
            )['total'] or 0
            
            # Caja estimada (ventas - premios pagados - comisiones pagadas)
            comisiones_pagadas_hoy = Liquidacion.objects.filter(
                empresario=user,
                fecha_pago__date=hoy,
                estado='pagada'
            ).aggregate(total=Sum('comision_valor'))['total'] or 0
            caja_hoy = ventas_hoy - premios_pagados_hoy - comisiones_pagadas_hoy
            
            # Liquidaciones pendientes
            liquidaciones_pendientes = Liquidacion.objects.filter(
                empresario=user,
                estado='solicitada'
            ).count()
            valor_liquidaciones_pendientes = Liquidacion.objects.filter(
                empresario=user,
                estado='solicitada'
            ).aggregate(total=Sum('comision_valor'))['total'] or 0
            
            # Desglose por chancero (todos los chanceros con sus ventas de hoy)
            chanceros_data = Usuario.objects.filter(
                rol='chancero',
                empresario=user
            ).annotate(
                ventas_hoy=Sum(
                    'apuestas_chancero__monto_apostado',
                    filter=Q(apuestas_chancero__fecha_hora__date=hoy)
                ),
                apuestas_hoy=Count(
                    'apuestas_chancero',
                    filter=Q(apuestas_chancero__fecha_hora__date=hoy)
                )
            ).order_by('-ventas_hoy')

            # Detalle de apuestas por chancero para mostrar números, loterías y horas
            apuestas_detalle_por_chancero = {}
            for chancero in chanceros_data:
                apuestas_chancero = Apuesta.objects.filter(
                    empresario=user,
                    chancero=chancero,
                    fecha_hora__date=hoy
                ).select_related('loteria').order_by('-fecha_hora')
                apuestas_detalle_por_chancero[chancero.id] = apuestas_chancero
            
            # Top 5 chanceros del día
            top_chanceros = chanceros_data[:5]
            
            # Loterías activas
            loterias_activas = Loteria.objects.filter(empresario=user, activa=True).count()
            
            # Loterías activas del día actual
            hoy_str = hoy.strftime('%A')
            dias_map = {
                'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
                'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado',
                'Sunday': 'Domingo'
            }
            dia_hoy = dias_map[hoy_str]
            
            loterias_hoy = Loteria.objects.filter(
                empresario=user,
                activa=True,
                dias_habilitados__contains=[dia_hoy]
            )
            
            # Ventas por lotería hoy
            ventas_por_loteria = Loteria.objects.filter(
                empresario=user,
                activa=True
            ).annotate(
                ventas_hoy=Sum(
                    'apuestas__monto_apostado',
                    filter=Q(apuestas__fecha_hora__date=hoy)
                )
            ).order_by('-ventas_hoy')
            
            context = {
                'total_chanceros': total_chanceros,
                'chanceros_activos': chanceros_activos,
                'ventas_hoy': ventas_hoy,
                'apuestas_count': apuestas_count,
                'premios_pagados_hoy': premios_pagados_hoy,
                'caja_hoy': caja_hoy,
                'liquidaciones_pendientes': liquidaciones_pendientes,
                'valor_liquidaciones_pendientes': valor_liquidaciones_pendientes,
                'top_chanceros': top_chanceros,
                'chanceros_data': chanceros_data,
                'apuestas_detalle_por_chancero': apuestas_detalle_por_chancero,
                'loterias_activas': loterias_activas,
                'loterias_hoy': loterias_hoy,
                'ventas_por_loteria': ventas_por_loteria,
            }
            return render(request, 'empresario_dashboard.html', context)
        
        elif user.rol == 'chancero':
            # Datos para chancero
            ahora_local = timezone.localtime(timezone.now())
            hoy = ahora_local.date()

            ventas_hoy = Apuesta.objects.filter(
                chancero=user,
                fecha_hora__date=hoy
            ).aggregate(total=Sum('monto_apostado'))['total'] or 0

            apuestas_hoy = Apuesta.objects.filter(
                chancero=user,
                fecha_hora__date=hoy
            ).count()

            try:
                comision = ComisionVendedor.objects.get(chancero=user)
                porcentaje = comision.porcentaje
            except:
                porcentaje = 0

            # Últimas 10 apuestas
            ultimas_apuestas = Apuesta.objects.filter(
                chancero=user
            ).order_by('-fecha_hora')[:10]

            # Loterías disponibles hoy (usar mismo filtrado que ApuestaCreateView)
            ahora_dt = timezone.localtime(timezone.now())
            hoy = ahora_dt.strftime('%A')
            dias_map = {
                'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
                'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado',
                'Sunday': 'Domingo'
            }
            dia_hoy = dias_map[hoy]

            loterias = Loteria.objects.filter(
                empresario=user.empresario,
                activa=True
            )

            loterias_disponibles = []
            hora_actual = ahora_dt.time()
            for loteria in loterias:
                dias = loteria.dias_habilitados or []
                if isinstance(dias, str):
                    dias = [dias]
                dia_habilitado = dia_hoy in dias
                horario_valido = loteria.hora_apertura <= hora_actual <= loteria.hora_cierre
                if dia_habilitado and horario_valido:
                    loterias_disponibles.append(loteria)

            context = {
                'ventas_hoy': ventas_hoy,
                'apuestas_hoy': apuestas_hoy,
                'porcentaje_comision': porcentaje,
                'ultimas_apuestas': ultimas_apuestas,
                'loterias_disponibles': loterias_disponibles,
            }
            return render(request, 'chancero_dashboard.html', context)


# ========== VISTAS PARA ADMIN GENERAL ==========

@method_decorator(login_required, name='dispatch')
class EmpresariosListView(View):
    def get(self, request):
        if request.user.rol != 'admin_general':
            return redirect('dashboard')
        
        empresarios = Usuario.objects.filter(rol='empresario').order_by('-created_at')
        return render(request, 'admin/empresarios.html', {'empresarios': empresarios})


@method_decorator(login_required, name='dispatch')
class EmpresarioCreateView(View):
    def get(self, request):
        if request.user.rol != 'admin_general':
            return redirect('dashboard')
        return render(request, 'admin/empresario_create.html')
    
    def post(self, request):
        if request.user.rol != 'admin_general':
            return redirect('dashboard')
        
        data = request.POST
        password = data.get('password')
        
        usuario = Usuario(
            documento=data.get('documento'),
            nombres=data.get('nombres'),
            apellidos=data.get('apellidos'),
            direccion=data.get('direccion'),
            telefono=data.get('telefono'),
            email=data.get('email', ''),
            telefono_familiar=data.get('telefono_familiar'),
            observaciones=data.get('observaciones', ''),
            rol='empresario',
            activo=True,
            fecha_afiliacion=data.get('fecha_afiliacion') or None,
            fecha_vencimiento=data.get('fecha_vencimiento') or None,
        )
        usuario.set_password(password)
        usuario.save()
        
        return redirect('dashboard')


@method_decorator(login_required, name='dispatch')
class EmpresarioDetailView(View):
    def get(self, request, pk):
        if request.user.rol != 'admin_general':
            return redirect('dashboard')
        
        empresario = get_object_or_404(Usuario, pk=pk, rol='empresario')
        
        # Chanceros del empresario
        chanceros = Usuario.objects.filter(empresario=empresario, rol='chancero')
        total_chanceros = chanceros.count()
        
        context = {
            'empresario': empresario,
            'chanceros': chanceros,
            'total_chanceros': total_chanceros,
        }
        return render(request, 'admin/empresario_detail.html', context)


@method_decorator(login_required, name='dispatch')
class EmpresarioUpdateView(View):
    def get(self, request, pk):
        if request.user.rol != 'admin_general':
            return redirect('dashboard')
        
        empresario = get_object_or_404(Usuario, pk=pk, rol='empresario')
        
        context = {
            'empresario': empresario,
        }
        return render(request, 'admin/empresario_form.html', context)
    
    def post(self, request, pk):
        if request.user.rol != 'admin_general':
            return redirect('dashboard')
        
        empresario = get_object_or_404(Usuario, pk=pk, rol='empresario')
        
        data = request.POST
        empresario.documento = data.get('documento')
        empresario.nombres = data.get('nombres')
        empresario.apellidos = data.get('apellidos')
        empresario.email = data.get('email', '')
        empresario.telefono = data.get('telefono')
        empresario.direccion = data.get('direccion')
        empresario.limite_chanceros = int(data.get('limite_chanceros', 10))
        empresario.telefono_familiar = data.get('telefono_familiar')
        empresario.activo = data.get('activo') == 'on'
        
        # Si se proporciona nueva contraseña, actualizarla
        nueva_password = data.get('password')
        if nueva_password:
            empresario.set_password(nueva_password)
        
        empresario.save()
        
        return redirect('empresario_detail', pk=empresario.id)


@method_decorator(login_required, name='dispatch')
class EmpresarioToggleActivoView(View):
    def post(self, request, pk):
        if request.user.rol != 'admin_general':
            return redirect('dashboard')
        
        empresario = get_object_or_404(Usuario, pk=pk, rol='empresario')
        empresario.activo = not empresario.activo
        empresario.save()
        
        return redirect('dashboard')


@method_decorator(login_required, name='dispatch')
class ComisionGlobalListView(View):
    def get(self, request):
        if request.user.rol != 'admin_general':
            return redirect('dashboard')
        
        comisiones = ComisionGlobal.objects.filter(activo=True).select_related('empresario')
        return render(request, 'admin/comisiones_global.html', {'comisiones': comisiones})


@method_decorator(login_required, name='dispatch')
class ComisionGlobalCreateView(View):
    def post(self, request):
        if request.user.rol != 'admin_general':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        empresario_id = request.POST.get('empresario_id')
        porcentaje = request.POST.get('porcentaje')
        
        # Desactivar comisión anterior si existe
        ComisionGlobal.objects.filter(empresario_id=empresario_id, activo=True).update(activo=False)
        
        comision = ComisionGlobal(
            empresario_id=empresario_id,
            porcentaje=porcentaje,
            activo=True
        )
        comision.save()
        
        return JsonResponse({'success': True, 'id': comision.id})


# ========== VISTAS PARA EMPRESARIO ==========

@method_decorator(login_required, name='dispatch')
class ChancerosListView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        chanceros = Usuario.objects.filter(
            rol='chancero', 
            empresario=request.user
        ).order_by('-created_at')
        
        # Obtener comisiones para cada chancero
        chanceros_con_comision = []
        for chancero in chanceros:
            try:
                comision = ComisionVendedor.objects.get(
                    empresario=request.user,
                    chancero=chancero
                )
                porcentaje = comision.porcentaje
            except ComisionVendedor.DoesNotExist:
                porcentaje = 0
            
            chanceros_con_comision.append({
                'chancero': chancero,
                'porcentaje_comision': porcentaje
            })
        
        return render(request, 'empresario/chanceros.html', {'chanceros_data': chanceros_con_comision})


@method_decorator(login_required, name='dispatch')
class ChanceroCreateView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        # Validar límite de chanceros antes de mostrar el formulario
        chanceros_actuales = Usuario.objects.filter(
            rol='chancero',
            empresario=request.user
        ).count()
        
        limite = request.user.limite_chanceros or 10
        if chanceros_actuales >= limite:
            return render(request, 'empresario/chancero_form.html', {
                'error': f'Has alcanzado el límite de {limite} chanceros. Contacta al administrador.'
            })
        
        return render(request, 'empresario/chancero_form.html')
    
    def post(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        # Validar límite de chanceros
        chanceros_actuales = Usuario.objects.filter(
            rol='chancero',
            empresario=request.user
        ).count()
        
        limite = request.user.limite_chanceros or 10
        if chanceros_actuales >= limite:
            return render(request, 'empresario/chancero_form.html', {
                'error': f'Has alcanzado el límite de {limite} chanceros. Contacta al administrador.'
            })
        
        data = request.POST
        password = data.get('password')
        
        usuario = Usuario(
            documento=data.get('documento'),
            nombres=data.get('nombres'),
            apellidos=data.get('apellidos'),
            direccion=data.get('direccion'),
            telefono=data.get('telefono'),
            email=data.get('email', ''),
            telefono_familiar=data.get('telefono_familiar'),
            observaciones=data.get('observaciones', ''),
            rol='chancero',
            empresario=request.user,
            activo=False,
        )
        usuario.set_password(password)
        usuario.save()
        
        # Guardar comisión
        comision = data.get('comision')
        if comision:
            ComisionVendedor.objects.create(
                empresario=request.user,
                chancero=usuario,
                porcentaje=comision
            )
        
        return redirect('chanceros_list')


@method_decorator(login_required, name='dispatch')
class ChanceroDetailView(View):
    def get(self, request, pk):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        chancero = get_object_or_404(Usuario, pk=pk, rol='chancero', empresario=request.user)
        
        # Estadísticas
        ventas_hoy = Apuesta.objects.filter(
            chancero=chancero,
            fecha_hora__date=timezone.now().date()
        ).aggregate(total=Sum('monto_apostado'))['total'] or 0
        
        ventas_semana = Apuesta.objects.filter(
            chancero=chancero,
            fecha_hora__gte=timezone.now() - timedelta(days=7)
        ).aggregate(total=Sum('monto_apostado'))['total'] or 0
        
        ultimas_apuestas = Apuesta.objects.filter(chancero=chancero).order_by('-fecha_hora')[:20]
        
        # Obtener comisión del chancero
        try:
            comision_obj = ComisionVendedor.objects.get(chancero=chancero, empresario=request.user)
            porcentaje_comision = float(comision_obj.porcentaje)
        except ComisionVendedor.DoesNotExist:
            porcentaje_comision = 0.0
        
        context = {
            'chancero': chancero,
            'ventas_hoy': ventas_hoy,
            'ventas_semana': ventas_semana,
            'ultimas_apuestas': ultimas_apuestas,
            'porcentaje_comision': porcentaje_comision,
        }
        return render(request, 'empresario/chancero_detail.html', context)


@method_decorator(login_required, name='dispatch')
class ChanceroToggleActivoView(View):
    def post(self, request, pk):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        chancero = get_object_or_404(Usuario, pk=pk, rol='chancero', empresario=request.user)
        chancero.activo = not chancero.activo
        chancero.save()
        
        return redirect('chancero_detail', pk=pk)


@method_decorator(login_required, name='dispatch')
class ChanceroUpdateView(View):
    def post(self, request, pk):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        chancero = get_object_or_404(Usuario, pk=pk, rol='chancero', empresario=request.user)
        
        data = request.POST
        chancero.nombres = data.get('nombres')
        chancero.apellidos = data.get('apellidos')
        chancero.telefono = data.get('telefono')
        chancero.direccion = data.get('direccion')
        chancero.email = data.get('email', '')
        chancero.telefono_familiar = data.get('telefono_familiar')
        
        # Actualizar contraseña si se proporciona
        nueva_password = data.get('password')
        if nueva_password:
            chancero.set_password(nueva_password)
        
        chancero.save()
        
        # Actualizar o crear comisión
        comision = data.get('comision')
        if comision:
            ComisionVendedor.objects.update_or_create(
                empresario=request.user,
                chancero=chancero,
                defaults={'porcentaje': comision}
            )
        
        return redirect('chancero_detail', pk=pk)


@method_decorator(login_required, name='dispatch')
class ComisionVendedorUpdateView(View):
    def post(self, request, pk):
        if request.user.rol != 'empresario':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        chancero = get_object_or_404(Usuario, pk=pk, rol='chancero', empresario=request.user)
        porcentaje = request.POST.get('porcentaje')
        
        comision, created = ComisionVendedor.objects.update_or_create(
            empresario=request.user,
            chancero=chancero,
            defaults={'porcentaje': porcentaje}
        )
        
        return JsonResponse({'success': True, 'porcentaje': float(porcentaje)})


@method_decorator(login_required, name='dispatch')
class LoteriasListView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        loterias = Loteria.objects.filter(empresario=request.user).order_by('nombre')
        return render(request, 'empresario/loterias.html', {'loterias': loterias})


@method_decorator(login_required, name='dispatch')
class LoteriaCreateView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        return render(request, 'empresario/loteria_form.html')
    
    def post(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        data = request.POST
        dias = data.getlist('dias')
        
        loteria = Loteria(
            empresario=request.user,
            nombre=data.get('nombre'),
            dias_habilitados=dias,
            hora_apertura=data.get('horario_apertura'),
            hora_cierre=data.get('horario_cierre'),
            tope_premio=data.get('tope_premio', 0),
            activa=data.get('activa') == 'on',
        )
        loteria.save()
        
        return redirect('loterias_list')


@method_decorator(login_required, name='dispatch')
class LoteriaToggleActivoView(View):
    def post(self, request, pk):
        if request.user.rol != 'empresario':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        loteria = get_object_or_404(Loteria, pk=pk, empresario=request.user)
        loteria.activa = not loteria.activa
        loteria.save()
        
        return JsonResponse({'activa': loteria.activa})


@method_decorator(login_required, name='dispatch')
class PlanesPremioListView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        planes = PlanPremio.objects.filter(empresario=request.user).order_by('cifras')
        
        context = {
            'planes': planes,
        }
        return render(request, 'empresario/planes_premio.html', context)


@method_decorator(login_required, name='dispatch')
class PlanesPremioAPIView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        planes = PlanPremio.objects.filter(empresario=request.user).order_by('cifras')
        
        planes_data = []
        for plan in planes:
            planes_data.append({
                'id': plan.id,
                'cifras': plan.cifras,
                'es_combinado': plan.es_combinado,
                'premio_por_peso': float(plan.premio_por_peso)
            })
        
        return JsonResponse(planes_data, safe=False)


@method_decorator(login_required, name='dispatch')
class ValidarApuestaAPIView(View):
    def post(self, request):
        if request.user.rol != 'chancero':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        data = json.loads(request.body)
        loteria_id = data.get('loteria_id')
        tipo_apuesta = data.get('tipo_apuesta')
        numeros_str = data.get('numeros')
        monto = float(data.get('monto', 0))
        
        # Validar lotería
        try:
            loteria = Loteria.objects.get(id=loteria_id, empresario=request.user.empresario, activa=True)
        except Loteria.DoesNotExist:
            return JsonResponse({'valido': False, 'error': 'Lotería no válida o inactiva'})
        
        # Validar números
        numeros_array = [n.strip() for n in numeros_str.split(',') if n.strip()]
        if not numeros_array:
            return JsonResponse({'valido': False, 'error': 'Debe ingresar al menos un número'})
        
        # Validar longitud de números
        longitud = len(numeros_array[0])
        if longitud < 2 or longitud > 6:
            return JsonResponse({'valido': False, 'error': 'Los números deben tener entre 2 y 6 cifras'})
        
        # Validar que todos los números tengan la misma longitud
        for num in numeros_array:
            if len(num) != longitud:
                return JsonResponse({'valido': False, 'error': 'Todos los números deben tener la misma cantidad de cifras'})
        
        # Validar autorización de cifras según configuración del empresario
        empresario = request.user.empresario
        es_combinado = tipo_apuesta == 'combinado'
        
        if longitud == 2:
            if es_combinado and not empresario.autorizar_2c_combinado:
                return JsonResponse({'valido': False, 'error': 'El empresario no autoriza apuestas de 2 cifras combinado'})
            if not es_combinado and not empresario.autorizar_2c_directo:
                return JsonResponse({'valido': False, 'error': 'El empresario no autoriza apuestas de 2 cifras directo'})
        elif longitud == 3:
            if es_combinado and not empresario.autorizar_3c_combinado:
                return JsonResponse({'valido': False, 'error': 'El empresario no autoriza apuestas de 3 cifras combinado'})
            if not es_combinado and not empresario.autorizar_3c_directo:
                return JsonResponse({'valido': False, 'error': 'El empresario no autoriza apuestas de 3 cifras directo'})
        elif longitud == 4:
            if es_combinado and not empresario.autorizar_4c_combinado:
                return JsonResponse({'valido': False, 'error': 'El empresario no autoriza apuestas de 4 cifras combinado'})
            if not es_combinado and not empresario.autorizar_4c_directo:
                return JsonResponse({'valido': False, 'error': 'El empresario no autoriza apuestas de 4 cifras directo'})
        elif longitud == 5:
            if es_combinado and not empresario.autorizar_5c_combinado:
                return JsonResponse({'valido': False, 'error': 'El empresario no autoriza apuestas de 5 cifras combinado'})
            if not es_combinado and not empresario.autorizar_5c_directo:
                return JsonResponse({'valido': False, 'error': 'El empresario no autoriza apuestas de 5 cifras directo'})
        elif longitud == 6:
            if es_combinado and not empresario.autorizar_6c_combinado:
                return JsonResponse({'valido': False, 'error': 'El empresario no autoriza apuestas de 6 cifras combinado'})
            if not es_combinado and not empresario.autorizar_6c_directo:
                return JsonResponse({'valido': False, 'error': 'El empresario no autoriza apuestas de 6 cifras directo'})
        
        # Calcular premio potencial
        try:
            plan = PlanPremio.objects.get(
                empresario=empresario,
                cifras=longitud,
                es_combinado=es_combinado
            )
            premio_por_peso = float(plan.premio_por_peso)
        except PlanPremio.DoesNotExist:
            return JsonResponse({'valido': False, 'error': f'No hay plan de premio configurado para {longitud} cifras {tipo_apuesta}'})
        
        # Calcular premio total
        if es_combinado:
            # Para combinado, se multiplican las combinaciones
            from itertools import permutations
            combinaciones = len(list(permutations(numeros_array[0], longitud)))
            premio_total = monto * premio_por_peso * combinaciones
        else:
            # Para directo, premio por número
            premio_total = monto * premio_por_peso * len(numeros_array)
        
        # Validar tope de premio de la lotería
        if premio_total > loteria.tope_premio:
            return JsonResponse({
                'valido': False,
                'error': f'El premio potencial (${premio_total:.0f}) excede el tope máximo de la lotería (${loteria.tope_premio:.0f})'
            })
        
        return JsonResponse({'valido': True})


@method_decorator(login_required, name='dispatch')
class LoteriasPorFechaView(View):
    def get(self, request):
        if request.user.rol != 'chancero':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        fecha = request.GET.get('fecha')
        if not fecha:
            return JsonResponse({'error': 'Debe especificar fecha'}, status=400)
        
        # Convertir fecha a día de la semana
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
        dias_semana = {
            0: 'Monday', 1: 'Tuesday', 2: 'Wednesday', 3: 'Thursday',
            4: 'Friday', 5: 'Saturday', 6: 'Sunday'
        }
        dia_semana = dias_semana[fecha_obj.weekday()]
        
        dias_map = {
            'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
            'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado',
            'Sunday': 'Domingo'
        }
        dia_hoy = dias_map[dia_semana]
        
        # Obtener loterías disponibles para esa fecha (case-insensitive)
        loterias = Loteria.objects.filter(
            empresario=request.user.empresario,
            activa=True
        )
        
        # Filtrar por día (case-insensitive)
        loterias_filtradas = []
        for loteria in loterias:
            dias_habilitados = loteria.dias_habilitados or []
            # Comparar en minúsculas para ser case-insensitive
            if any(dia.lower() == dia_hoy.lower() for dia in dias_habilitados):
                loterias_filtradas.append(loteria)
        
        loterias_data = []
        for loteria in loterias_filtradas:
            loterias_data.append({
                'id': loteria.id,
                'nombre': loteria.nombre,
                'tope_premio': float(loteria.tope_premio)
            })
        
        return JsonResponse({'loterias': loterias_data})


@method_decorator(login_required, name='dispatch')
class CalcularPremioAPIView(View):
    def post(self, request):
        if request.user.rol != 'chancero':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        data = json.loads(request.body)
        loteria_id = data.get('loteria_id')
        apuestas = data.get('apuestas', [])
        
        # Obtener lotería
        try:
            loteria = Loteria.objects.get(id=loteria_id, empresario=request.user.empresario)
        except Loteria.DoesNotExist:
            return JsonResponse({'error': 'Lotería no válida'})
        
        premio_total = 0
        premio_ultima_apuesta = 0
        excede_tope = False
        maximo_permitido = 0
        from itertools import permutations
        acumulados_temporales = {}

        for apuesta in apuestas:
            numero = apuesta.get('numero')
            cifras = apuesta.get('cifras')
            monto = float(apuesta.get('monto', 0))
            es_combinado = apuesta.get('es_combinado', False)

            if not numero or not cifras or not monto:
                continue

            try:
                plan = PlanPremio.objects.get(
                    empresario=request.user.empresario,
                    cifras=cifras,
                    es_combinado=es_combinado
                )
                premio_por_peso = float(plan.premio_por_peso)
            except PlanPremio.DoesNotExist:
                return JsonResponse({'error': f'No hay plan de premio configurado para {cifras} cifras tipo {"combinado" if es_combinado else "pleno"}'})

            if es_combinado:
                numero_generico = '1' * cifras
                combinaciones = len(list(permutations(numero_generico, cifras)))
                divisor_tope = premio_por_peso * combinaciones
            else:
                combinaciones = 1
                divisor_tope = premio_por_peso

            clave = f'{loteria.id}-{numero}-{cifras}-{es_combinado}'
            if clave not in acumulados_temporales:
                monto_guardado = Apuesta.objects.filter(
                    empresario=request.user.empresario,
                    loteria=loteria,
                    numero=numero,
                    cifras=cifras,
                    estado='activa'
                ).aggregate(total=Sum('monto_apostado'))['total'] or 0
                acumulados_temporales[clave] = float(monto_guardado)

            monto_acumulado = acumulados_temporales[clave]
            premio_apuesta = monto * premio_por_peso * combinaciones
            premio_ultima_apuesta = premio_apuesta
            premio_potencial_total = (monto_acumulado + monto) * divisor_tope
            premio_total += premio_apuesta

            if premio_potencial_total > float(loteria.tope_premio):
                excede_tope = True
                tope_premio_float = float(loteria.tope_premio)
                maximo_permitido = max((tope_premio_float / divisor_tope) - monto_acumulado, 0)
                break

            acumulados_temporales[clave] += monto

        return JsonResponse({
            'premio_potencial': premio_total,
            'premio_apuesta': premio_ultima_apuesta,
            'excede_tope': excede_tope,
            'maximo_permitido': maximo_permitido
        })


@method_decorator(login_required, name='dispatch')
class RealizarApuestaAPIView(View):
    def post(self, request):
        if request.user.rol != 'chancero':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        data = json.loads(request.body)
        loteria_id = data.get('loteria_id')
        fecha = data.get('fecha')
        apuestas = data.get('apuestas', [])
        telefono_cliente = data.get('telefono_cliente', '')
        
        # Validar lotería
        try:
            loteria = Loteria.objects.get(id=loteria_id, empresario=request.user.empresario, activa=True)
        except Loteria.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Lotería no válida o inactiva'})
        
        from itertools import permutations
        
        apuestas_creadas = []
        total_apuestas = 0
        
        # Procesar cada apuesta individualmente
        for apuesta_data in apuestas:
            numero = apuesta_data.get('numero')
            cifras = apuesta_data.get('cifras')
            monto = float(apuesta_data.get('monto', 0))
            es_combinado = apuesta_data.get('es_combinado', False)
            
            if not numero or not cifras or not monto:
                continue
            
            # Obtener plan de premio
            try:
                plan = PlanPremio.objects.get(
                    empresario=request.user.empresario,
                    cifras=cifras,
                    es_combinado=es_combinado
                )
                premio_por_peso = float(plan.premio_por_peso)
            except PlanPremio.DoesNotExist:
                return JsonResponse({'success': False, 'error': f'No hay plan de premio configurado para {cifras} cifras tipo {"combinado" if es_combinado else "pleno"}'})
            
            # Calcular premio potencial
            if es_combinado:
                numero_generico = '1' * cifras
                combinaciones = len(list(permutations(numero_generico, cifras)))
                premio_potencial = monto * premio_por_peso * combinaciones
            else:
                premio_potencial = monto * premio_por_peso
            
            monto_acumulado = float(Apuesta.objects.filter(
                empresario=request.user.empresario,
                loteria=loteria,
                numero=numero,
                cifras=cifras,
                estado='activa'
            ).aggregate(total=Sum('monto_apostado'))['total'] or 0)
            
            # Validar que no exceda el tope máximo de la lotería
            if es_combinado:
                premio_potencial_total = (monto_acumulado + monto) * premio_por_peso * combinaciones
            else:
                premio_potencial_total = (monto_acumulado + monto) * premio_por_peso
            
            if premio_potencial_total > float(loteria.tope_premio):
                return JsonResponse({
                    'success': False,
                    'error': f'El premio potencial (${premio_potencial_total:.0f}) excede el tope máximo de la lotería (${loteria.tope_premio:.0f}). Acumulado actual: ${monto_acumulado:.0f}. Puede que otra apuesta se haya realizado simultáneamente.'
                })
            
            # Crear apuesta individual con la fecha especificada
            fecha_hora_apuesta = timezone.now()
            if fecha:
                fecha_hora_apuesta = datetime.strptime(fecha, '%Y-%m-%d').replace(
                    hour=timezone.now().hour,
                    minute=timezone.now().minute,
                    second=timezone.now().second
                )
            
            apuesta = Apuesta(
                empresario=request.user.empresario,
                chancero=request.user,
                loteria=loteria,
                numero=numero,
                cifras=cifras,
                monto_apostado=monto,
                premio_potencial=premio_potencial,
                estado='activa',
                fecha_hora=fecha_hora_apuesta
            )
            apuesta.save()
            apuestas_creadas.append(apuesta)
            total_apuestas += 1

            acumulado, created = AcumuladoNumero.objects.get_or_create(
                empresario=request.user.empresario,
                loteria=loteria,
                numero=numero,
                cifras=cifras,
                defaults={'monto_acumulado': 0}
            )
            acumulado.monto_acumulado = Apuesta.objects.filter(
                empresario=request.user.empresario,
                loteria=loteria,
                numero=numero,
                cifras=cifras,
                estado='activa'
            ).aggregate(total=Sum('monto_apostado'))['total'] or 0
            acumulado.fecha_ultima_apuesta = timezone.now()
            acumulado.save()
            
            # Crear historial de cliente si se proporcionó teléfono
            if telefono_cliente:
                try:
                    cliente = Cliente.objects.get(
                        chancero=request.user,
                        telefono=telefono_cliente,
                        activo=True
                    )
                    HistorialCliente.objects.create(
                        cliente=cliente,
                        apuesta=apuesta
                    )
                except Cliente.DoesNotExist:
                    # No hacer nada si el cliente no existe
                    pass
        
        return JsonResponse({
            'success': True,
            'total_apuestas': total_apuestas
        })


@method_decorator(login_required, name='dispatch')
class PlanPremioUpdateView(View):
    def post(self, request):
        if request.user.rol != 'empresario':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        cifras = request.POST.get('cifras')
        es_combinado = request.POST.get('es_combinado') == 'true'
        premio_por_peso = request.POST.get('premio_por_peso')
        
        plan, created = PlanPremio.objects.update_or_create(
            empresario=request.user,
            cifras=cifras,
            es_combinado=es_combinado,
            defaults={'premio_por_peso': premio_por_peso}
        )
        
        return JsonResponse({'success': True})


@method_decorator(login_required, name='dispatch')
class TopesListView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        loteria_id = request.GET.get('loteria')
        if loteria_id:
            topes = TopeNumero.objects.filter(
                empresario=request.user,
                loteria_id=loteria_id
            ).order_by('numero')
        else:
            topes = TopeNumero.objects.filter(empresario=request.user).order_by('loteria', 'numero')
        
        loterias = Loteria.objects.filter(empresario=request.user, activa=True)
        
        context = {
            'topes': topes,
            'loterias': loterias,
            'loteria_seleccionada': loteria_id,
        }
        return render(request, 'empresario/topes.html', context)


@method_decorator(login_required, name='dispatch')
class TopeCreateView(View):
    def post(self, request):
        if request.user.rol != 'empresario':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        loteria_id = request.POST.get('loteria_id')
        numero = request.POST.get('numero')
        tope = request.POST.get('tope')
        
        tope_obj, created = TopeNumero.objects.update_or_create(
            empresario=request.user,
            loteria_id=loteria_id,
            numero=numero,
            defaults={'tope_maximo': tope}
        )
        
        return JsonResponse({'success': True, 'id': tope_obj.id})


@method_decorator(login_required, name='dispatch')
class VentasListView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        fecha_inicio = request.GET.get('fecha_inicio', (timezone.now() - timedelta(days=7)).strftime('%Y-%m-%d'))
        fecha_fin = request.GET.get('fecha_fin', timezone.now().strftime('%Y-%m-%d'))
        chancero_id = request.GET.get('chancero')
        loteria_id = request.GET.get('loteria')
        
        apuestas = Apuesta.objects.filter(
            empresario=request.user,
            fecha_hora__date__gte=fecha_inicio,
            fecha_hora__date__lte=fecha_fin
        )
        
        if chancero_id:
            apuestas = apuestas.filter(chancero_id=chancero_id)
        if loteria_id:
            apuestas = apuestas.filter(loteria_id=loteria_id)
        
        apuestas = apuestas.select_related('chancero', 'loteria').order_by('-fecha_hora')
        
        total_ventas = apuestas.aggregate(total=Sum('monto_apostado'))['total'] or 0
        total_premios = apuestas.filter(estado='pagada').aggregate(total=Sum('premio_potencial'))['total'] or 0
        ganancia_neta = total_ventas - total_premios
        
        chanceros = Usuario.objects.filter(rol='chancero', empresario=request.user)
        loterias = Loteria.objects.filter(empresario=request.user)
        
        context = {
            'apuestas': apuestas,
            'total_ventas': total_ventas,
            'total_premios': total_premios,
            'ganancia_neta': ganancia_neta,
            'chanceros': chanceros,
            'loterias': loterias,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'chancero_id': chancero_id,
            'loteria_id': loteria_id,
        }
        return render(request, 'empresario/ventas.html', context)


@method_decorator(login_required, name='dispatch')
class ReportesChanceroView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        chancero_id = request.GET.get('chancero')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        chanceros = Usuario.objects.filter(rol='chancero', empresario=request.user)
        
        apuestas = []
        total_ventas = 0
        total_premios = 0
        total_apuestas = 0
        total_comision = 0
        porcentaje_comision = 0
        
        if chancero_id and fecha_inicio and fecha_fin:
            apuestas = Apuesta.objects.filter(
                empresario=request.user,
                chancero_id=chancero_id,
                fecha_hora__date__gte=fecha_inicio,
                fecha_hora__date__lte=fecha_fin
            ).select_related('chancero', 'loteria').order_by('-fecha_hora')
            
            total_ventas = apuestas.aggregate(total=Sum('monto_apostado'))['total'] or 0
            total_premios = apuestas.filter(estado='pagada').aggregate(total=Sum('premio_potencial'))['total'] or 0
            total_apuestas = apuestas.count()
            
            # Obtener porcentaje de comisión del chancero
            try:
                comision = ComisionVendedor.objects.get(chancero_id=chancero_id)
                porcentaje_comision = comision.porcentaje
                total_comision = total_ventas * (porcentaje_comision / 100)
            except ComisionVendedor.DoesNotExist:
                porcentaje_comision = 0
                total_comision = 0
        
        context = {
            'chanceros': chanceros,
            'chancero_id': chancero_id,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'apuestas': apuestas,
            'total_ventas': total_ventas,
            'total_premios': total_premios,
            'total_apuestas': total_apuestas,
            'total_comision': total_comision,
            'porcentaje_comision': porcentaje_comision,
        }
        return render(request, 'empresario/reportes_chancero.html', context)


@method_decorator(login_required, name='dispatch')
class GananciasView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        # Calcular ganancias por fecha
        ganancias = []
        
        if fecha_inicio and fecha_fin:
            # Ventas totales
            apuestas = Apuesta.objects.filter(
                empresario=request.user,
                fecha_hora__date__gte=fecha_inicio,
                fecha_hora__date__lte=fecha_fin
            )
            
            total_ventas = apuestas.aggregate(total=Sum('monto_apostado'))['total'] or 0
            total_premios_pagados = apuestas.filter(estado='pagada').aggregate(total=Sum('premio_potencial'))['total'] or 0
            total_premios_potenciales = apuestas.aggregate(total=Sum('premio_potencial'))['total'] or 0
            
            # Comisiones pagadas a chanceros (liquidaciones pagadas en el período)
            liquidaciones_pagadas = Liquidacion.objects.filter(
                empresario=request.user,
                estado='pagada',
                fecha_inicio__gte=fecha_inicio,
                fecha_fin__lte=fecha_fin
            )
            total_comisiones_pagadas = liquidaciones_pagadas.aggregate(total=Sum('comision_valor'))['total'] or 0
            
            # Ganancia neta del empresario (ventas - premios pagados - comisiones pagadas)
            ganancia_neta_empresario = total_ventas - total_premios_pagados - total_comisiones_pagadas
            
            # Ganancia por lotería
            ganancias_por_loteria = apuestas.values('loteria__nombre').annotate(
                ventas=Sum('monto_apostado'),
                premios_pagados=Sum('premio_potencial', filter=Q(estado='pagada')),
                premios_potenciales=Sum('premio_potencial')
            ).annotate(
                ganancia=F('ventas') - F('premios_pagados')
            )
            
            # Ganancia por chancero (ventas - premios pagados - comisión)
            ganancias_por_chancero = apuestas.values('chancero__nombres', 'chancero__apellidos', 'chancero__id').annotate(
                ventas=Sum('monto_apostado'),
                premios_pagados=Sum('premio_potencial', filter=Q(estado='pagada')),
                premios_potenciales=Sum('premio_potencial')
            )
            
            # Calcular comisión y ganancia neta por chancero
            for item in ganancias_por_chancero:
                try:
                    comision = ComisionVendedor.objects.get(chancero_id=item['chancero__id'])
                    porcentaje = comision.porcentaje
                    comision_valor = item['ventas'] * (porcentaje / 100)
                    item['comision'] = comision_valor
                    item['ganancia_neta'] = item['ventas'] - item['premios_pagados'] - comision_valor
                    item['porcentaje_comision'] = porcentaje
                except ComisionVendedor.DoesNotExist:
                    item['comision'] = 0
                    item['ganancia_neta'] = item['ventas'] - item['premios_pagados']
                    item['porcentaje_comision'] = 0
            
            # Detalle de premios pagados en el período
            pagos_premios = PagoPremio.objects.filter(
                apuesta__empresario=request.user,
                fecha_pago__date__gte=fecha_inicio,
                fecha_pago__date__lte=fecha_fin
            ).select_related('apuesta', 'apuesta__chancero', 'apuesta__loteria', 'pagado_por').order_by('-fecha_pago')
            
            context = {
                'fecha_inicio': fecha_inicio,
                'fecha_fin': fecha_fin,
                'total_ventas': total_ventas,
                'total_premios_pagados': total_premios_pagados,
                'total_premios_potenciales': total_premios_potenciales,
                'total_comisiones_pagadas': total_comisiones_pagadas,
                'ganancia_neta_empresario': ganancia_neta_empresario,
                'ganancias_por_loteria': ganancias_por_loteria,
                'ganancias_por_chancero': ganancias_por_chancero,
                'pagos_premios': pagos_premios,
            }
        else:
            context = {
                'fecha_inicio': fecha_inicio,
                'fecha_fin': fecha_fin,
            }
        
        return render(request, 'empresario/ganancias.html', context)


@method_decorator(login_required, name='dispatch')
class LiquidacionesChanceroView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')

        chancero_id = request.GET.get('chancero')
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')

        chanceros = Usuario.objects.filter(rol='chancero', empresario=request.user)

        liquidaciones = []
        total_liquidado = Decimal('0')

        if chancero_id and fecha_inicio and fecha_fin:
            liquidaciones = Liquidacion.objects.filter(
                empresario=request.user,
                chancero_id=chancero_id,
                fecha_inicio__gte=fecha_inicio,
                fecha_fin__lte=fecha_fin
            ).select_related('chancero').order_by('-fecha_solicitud')

            from decimal import Decimal
            total_liquidado = liquidaciones.aggregate(total=Sum('comision_valor'))['total'] or Decimal('0')

        context = {
            'chanceros': chanceros,
            'chancero_id': chancero_id,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'liquidaciones': liquidaciones,
            'total_liquidado': total_liquidado,
        }
        return render(request, 'empresario/liquidaciones_chancero.html', context)


@method_decorator(login_required, name='dispatch')
class ExportarDatosEmpresarioView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return JsonResponse({'error': 'Debe especificar fecha inicio y fin'}, status=400)
        
        import pandas as pd
        from django.http import HttpResponse
        
        # Obtener datos
        apuestas = Apuesta.objects.filter(
            empresario=request.user,
            fecha_hora__date__gte=fecha_inicio,
            fecha_hora__date__lte=fecha_fin
        ).select_related('chancero', 'loteria').order_by('-fecha_hora')
        
        # Crear DataFrame
        data = []
        for apuesta in apuestas:
            data.append({
                'Fecha': apuesta.fecha_hora.strftime('%Y-%m-%d'),
                'Hora': apuesta.fecha_hora.strftime('%H:%M'),
                'Chancero': f"{apuesta.chancero.nombres} {apuesta.chancero.apellidos}",
                'Lotería': apuesta.loteria.nombre,
                'Número': apuesta.numero,
                'Cifras': apuesta.cifras,
                'Monto Apostado': float(apuesta.monto_apostado),
                'Premio Potencial': float(apuesta.premio_potencial),
                'Estado': apuesta.estado,
            })
        
        df = pd.DataFrame(data)
        
        # Crear respuesta Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=backup_{fecha_inicio}_a_{fecha_fin}.xlsx'
        
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Apuestas', index=False)
            
            # Agregar hoja de resumen
            resumen = {
                'Total Ventas': [df['Monto Apostado'].sum()],
                'Total Premios Potenciales': [df['Premio Potencial'].sum()],
                'Total Apuestas': [len(df)],
            }
            df_resumen = pd.DataFrame(resumen)
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
        
        return response


@method_decorator(login_required, name='dispatch')
class ConfiguracionRetencionView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        context = {
            'dias_retencion': request.user.dias_retencion_datos,
        }
        return render(request, 'empresario/configuracion_retencion.html', context)
    
    def post(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        dias = int(request.POST.get('dias_retencion', 90))
        request.user.dias_retencion_datos = dias
        request.user.save()
        
        return redirect('configuracion_retencion')


@method_decorator(login_required, name='dispatch')
class LimpiarDatosView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        from datetime import timedelta
        fecha_limite = timezone.now() - timedelta(days=request.user.dias_retencion_datos)
        
        # Contar registros a eliminar
        apuestas = Apuesta.objects.filter(
            empresario=request.user,
            fecha_hora__lt=fecha_limite
        )
        total_apuestas = apuestas.count()
        
        acumulados = AcumuladoNumero.objects.filter(
            empresario=request.user,
            fecha_ultima_apuesta__lt=fecha_limite
        )
        total_acumulados = acumulados.count()
        
        liquidaciones = Liquidacion.objects.filter(
            empresario=request.user,
            fecha_inicio__lt=fecha_limite
        )
        total_liquidaciones = liquidaciones.count()
        
        context = {
            'dias_retencion': request.user.dias_retencion_datos,
            'fecha_limite': fecha_limite.strftime('%Y-%m-%d'),
            'total_apuestas': total_apuestas,
            'total_acumulados': total_acumulados,
            'total_liquidaciones': total_liquidaciones,
            'total_registros': total_apuestas + total_acumulados + total_liquidaciones,
        }
        return render(request, 'empresario/limpiar_datos.html', context)
    
    def post(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        from datetime import timedelta
        fecha_limite = timezone.now() - timedelta(days=request.user.dias_retencion_datos)
        
        # Eliminar datos
        apuestas_eliminadas = Apuesta.objects.filter(
            empresario=request.user,
            fecha_hora__lt=fecha_limite
        ).delete()[0]
        
        acumulados_eliminados = AcumuladoNumero.objects.filter(
            empresario=request.user,
            fecha_ultima_apuesta__lt=fecha_limite
        ).delete()[0]
        
        liquidaciones_eliminadas = Liquidacion.objects.filter(
            empresario=request.user,
            fecha_inicio__lt=fecha_limite
        ).delete()[0]
        
        context = {
            'apuestas_eliminadas': apuestas_eliminadas,
            'acumulados_eliminados': acumulados_eliminados,
            'liquidaciones_eliminadas': liquidaciones_eliminadas,
            'total_eliminado': apuestas_eliminadas + acumulados_eliminados + liquidaciones_eliminadas,
        }
        return render(request, 'empresario/limpiar_datos_confirmacion.html', context)


@method_decorator(login_required, name='dispatch')
class ExportarDatosChanceroView(View):
    def get(self, request):
        if request.user.rol != 'chancero':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return JsonResponse({'error': 'Debe especificar fecha inicio y fin'}, status=400)
        
        import pandas as pd
        from django.http import HttpResponse
        
        # Obtener datos del chancero
        apuestas = Apuesta.objects.filter(
            chancero=request.user,
            fecha_hora__date__gte=fecha_inicio,
            fecha_hora__date__lte=fecha_fin
        ).select_related('loteria').order_by('-fecha_hora')
        
        # Obtener comisión
        try:
            from core.models import ComisionVendedor
            comision = ComisionVendedor.objects.get(chancero=request.user)
            porcentaje_comision = comision.porcentaje
        except ComisionVendedor.DoesNotExist:
            porcentaje_comision = 0
        
        # Crear DataFrame
        data = []
        for apuesta in apuestas:
            data.append({
                'Fecha': apuesta.fecha_hora.strftime('%Y-%m-%d'),
                'Hora': apuesta.fecha_hora.strftime('%H:%M'),
                'Lotería': apuesta.loteria.nombre,
                'Número': apuesta.numero,
                'Cifras': apuesta.cifras,
                'Monto Apostado': float(apuesta.monto_apostado),
                'Premio Potencial': float(apuesta.premio_potencial),
                'Estado': apuesta.estado,
            })
        
        df = pd.DataFrame(data)
        
        # Crear respuesta Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=backup_chancero_{fecha_inicio}_a_{fecha_fin}.xlsx'
        
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Mis Apuestas', index=False)
            
            # Agregar hoja de resumen
            total_ventas = df['Monto Apostado'].sum()
            total_comision = total_ventas * (porcentaje_comision / 100)
            
            resumen = {
                'Total Ventas': [total_ventas],
                'Mi Ganancia (Comisión)': [total_comision],
                'Porcentaje Comisión': [porcentaje_comision],
                'Total Apuestas': [len(df)],
            }
            df_resumen = pd.DataFrame(resumen)
            df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
        
        return response


@method_decorator(login_required, name='dispatch')
class ReportesPersonalesChanceroView(View):
    def get(self, request):
        if request.user.rol != 'chancero':
            return redirect('dashboard')
        
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        
        apuestas = []
        total_ventas = 0
        total_premios_pagados = 0
        total_premios_potenciales = 0
        total_comision = 0
        porcentaje_comision = 0
        total_apuestas = 0
        
        if fecha_inicio and fecha_fin:
            apuestas = Apuesta.objects.filter(
                chancero=request.user,
                fecha_hora__date__gte=fecha_inicio,
                fecha_hora__date__lte=fecha_fin
            ).select_related('loteria').order_by('-fecha_hora')
            
            total_ventas = apuestas.aggregate(total=Sum('monto_apostado'))['total'] or 0
            total_premios_pagados = apuestas.filter(estado='pagada').aggregate(total=Sum('premio_potencial'))['total'] or 0
            total_premios_potenciales = apuestas.aggregate(total=Sum('premio_potencial'))['total'] or 0
            total_apuestas = apuestas.count()
            
            # Obtener porcentaje de comisión del chancero
            try:
                comision = ComisionVendedor.objects.get(chancero=request.user)
                porcentaje_comision = comision.porcentaje
                total_comision = total_ventas * (porcentaje_comision / 100)
            except ComisionVendedor.DoesNotExist:
                porcentaje_comision = 0
                total_comision = 0
        
        context = {
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'apuestas': apuestas,
            'total_ventas': total_ventas,
            'total_premios_pagados': total_premios_pagados,
            'total_premios_potenciales': total_premios_potenciales,
            'total_comision': total_comision,
            'porcentaje_comision': porcentaje_comision,
            'total_apuestas': total_apuestas,
        }
        return render(request, 'chancero/reportes_personales.html', context)


@method_decorator(login_required, name='dispatch')
class ConsultarPremioChanceroView(View):
    def get(self, request):
        if request.user.rol != 'chancero':
            return redirect('dashboard')
        
        numero = request.GET.get('numero', '')
        fecha = request.GET.get('fecha', timezone.now().strftime('%Y-%m-%d'))
        
        loterias = Loteria.objects.filter(empresario=request.user.empresario)
        apuestas = []
        total_ventas = 0
        total_premios = 0
        total_comision = 0
        porcentaje_comision = 0
        
        if numero and fecha:
            apuestas = Apuesta.objects.filter(
                chancero=request.user,
                numero=numero,
                fecha_hora__date=fecha
            ).select_related('loteria').order_by('-fecha_hora')
            
            total_ventas = apuestas.aggregate(total=Sum('monto_apostado'))['total'] or 0
            total_premios = apuestas.aggregate(total=Sum('premio_potencial'))['total'] or 0
            
            # Obtener porcentaje de comisión del chancero
            try:
                comision = ComisionVendedor.objects.get(chancero=request.user)
                porcentaje_comision = comision.porcentaje
                total_comision = total_ventas * (porcentaje_comision / 100)
            except ComisionVendedor.DoesNotExist:
                porcentaje_comision = 0
                total_comision = 0
            
            # Calcular premios ganados
            for apuesta in apuestas:
                try:
                    resultado = ResultadoLoteria.objects.get(
                        loteria=apuesta.loteria,
                        fecha=fecha
                    )
                    numero_ganador = resultado.numero_ganador
                    if apuesta.numero.endswith(numero_ganador[-apuesta.cifras:]):
                        apuesta.premio_ganado = apuesta.premio_potencial
                    else:
                        apuesta.premio_ganado = 0
                except ResultadoLoteria.DoesNotExist:
                    apuesta.premio_ganado = None
        
        context = {
            'numero': numero,
            'fecha': fecha,
            'loterias': loterias,
            'apuestas': apuestas,
            'total_ventas': total_ventas,
            'total_premios': total_premios,
            'total_comision': total_comision,
            'porcentaje_comision': porcentaje_comision,
        }
        return render(request, 'chancero/consultar_premio.html', context)


@method_decorator(login_required, name='dispatch')
class AcumuladosView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        # Obtener todos los acumulados del empresario
        acumulados = AcumuladoNumero.objects.filter(
            empresario=request.user
        ).order_by('-monto_acumulado')
        
        context = {
            'acumulados': acumulados,
        }
        return render(request, 'empresario/acumulados.html', context)


@method_decorator(login_required, name='dispatch')
class ConsultarNumeroView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        numero = request.GET.get('numero', '')
        loteria_id = request.GET.get('loteria')
        fecha = request.GET.get('fecha', timezone.now().strftime('%Y-%m-%d'))
        
        loterias = Loteria.objects.filter(empresario=request.user)
        apuestas = []
        total_ventas = 0
        total_premios = 0
        
        if numero and loteria_id:
            apuestas = Apuesta.objects.filter(
                empresario=request.user,
                loteria_id=loteria_id,
                numero=numero,
                fecha_hora__date=fecha
            ).select_related('chancero', 'loteria').order_by('-fecha_hora')
            
            total_ventas = apuestas.aggregate(total=Sum('monto_apostado'))['total'] or 0
            total_premios = apuestas.aggregate(total=Sum('premio_potencial'))['total'] or 0
            
            # Calcular premio a pagar por cada apuesta si el número ganó
            # Para esto necesitamos verificar si hay resultado para esa lotería y fecha
            try:
                resultado = ResultadoLoteria.objects.get(
                    loteria_id=loteria_id,
                    fecha=fecha
                )
                numero_ganador = resultado.numero_ganador
                
                # Calcular premios ganados por cada apuesta
                for apuesta in apuestas:
                    # Verificar si el número coincide con el ganador según las cifras
                    if apuesta.numero.endswith(numero_ganador[-apuesta.cifras:]):
                        apuesta.premio_ganado = apuesta.premio_potencial
                    else:
                        apuesta.premio_ganado = 0
            except ResultadoLoteria.DoesNotExist:
                # No hay resultado, no se pueden calcular premios
                for apuesta in apuestas:
                    apuesta.premio_ganado = None
        
        context = {
            'numero': numero,
            'loterias': loterias,
            'loteria_id': loteria_id,
            'fecha': fecha,
            'apuestas': apuestas,
            'total_ventas': total_ventas,
            'total_premios': total_premios,
        }
        return render(request, 'empresario/consultar_numero.html', context)


@method_decorator(login_required, name='dispatch')
class PagarPremioView(View):
    def post(self, request, apuesta_id):
        if request.user.rol != 'empresario':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        apuesta = get_object_or_404(Apuesta, pk=apuesta_id, empresario=request.user)
        
        if apuesta.estado == 'pagada':
            return JsonResponse({'error': 'Esta apuesta ya fue pagada'}, status=400)
        
        # Verificar si la liquidación está pagada (autoriza pago de premios)
        if not apuesta.liquidacion_pagada:
            return JsonResponse({'error': 'La liquidación de esta apuesta no ha sido pagada. Pague la liquidación primero.'}, status=400)
        
        # Crear registro de pago
        pago = PagoPremio.objects.create(
            apuesta=apuesta,
            monto_pagado=apuesta.premio_potencial,
            pagado_por=request.user,
            observaciones=f"Pago de premio para número {apuesta.numero} en lotería {apuesta.loteria.nombre}"
        )
        
        # Actualizar estado de la apuesta
        apuesta.estado = 'pagada'
        apuesta.save()
        
        return JsonResponse({
            'success': True,
            'pago_id': pago.id,
            'monto_pagado': float(pago.monto_pagado),
            'fecha_pago': pago.fecha_pago.strftime('%Y-%m-%d %H:%M')
        })


@method_decorator(login_required, name='dispatch')
class ResultadosListView(View):
    def get(self, request):
        if request.user.rol not in ['admin_general', 'empresario']:
            return redirect('dashboard')
        
        if request.user.rol == 'empresario':
            resultados = ResultadoLoteria.objects.filter(
                loteria__empresario=request.user
            ).select_related('loteria', 'creado_por').order_by('-fecha', '-created_at')
        else:
            resultados = ResultadoLoteria.objects.all().select_related('loteria', 'creado_por').order_by('-fecha', '-created_at')
        
        context = {
            'resultados': resultados,
        }
        return render(request, 'empresario/resultados.html', context)


@method_decorator(login_required, name='dispatch')
class ResultadoCreateView(View):
    def get(self, request):
        if request.user.rol not in ['admin_general', 'empresario']:
            return redirect('dashboard')
        
        if request.user.rol == 'empresario':
            loterias = Loteria.objects.filter(empresario=request.user, activa=True)
        else:
            loterias = Loteria.objects.filter(activa=True)
        
        context = {
            'loterias': loterias,
        }
        return render(request, 'empresario/resultado_form.html', context)
    
    def post(self, request):
        if request.user.rol not in ['admin_general', 'empresario']:
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        loteria_id = request.POST.get('loteria_id')
        fecha = request.POST.get('fecha')
        numero_ganador = request.POST.get('numero_ganador')
        
        loteria = get_object_or_404(Loteria, pk=loteria_id)
        
        if request.user.rol == 'empresario' and loteria.empresario != request.user:
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        # Verificar si ya existe el resultado
        if ResultadoLoteria.objects.filter(loteria=loteria, fecha=fecha, numero_ganador=numero_ganador).exists():
            return JsonResponse({'error': 'Este resultado ya existe'}, status=400)
        
        resultado = ResultadoLoteria(
            loteria=loteria,
            fecha=fecha,
            numero_ganador=numero_ganador,
            cifras=len(numero_ganador),
            creado_por=request.user
        )
        resultado.save()
        
        return JsonResponse({'success': True, 'id': resultado.id})


@method_decorator(login_required, name='dispatch')
class VerificarPremiosView(View):
    def get(self, request, pk):
        if request.user.rol not in ['admin_general', 'empresario']:
            return redirect('dashboard')
        
        resultado = get_object_or_404(ResultadoLoteria, pk=pk)
        
        if request.user.rol == 'empresario' and resultado.loteria.empresario != request.user:
            return redirect('dashboard')
        
        # Buscar apuestas ganadoras del día
        apuestas_ganadoras = Apuesta.objects.filter(
            loteria=resultado.loteria,
            fecha_hora__date=resultado.fecha,
            numero__endswith=resultado.numero_ganador[-resultado.cifras:],
            estado='activa',
            liquidacion_pagada=True  # Solo apuestas liquidadas pueden pagar premios
        ).select_related('chancero')
        
        total_premios = apuestas_ganadoras.aggregate(total=Sum('premio_potencial'))['total'] or 0
        
        context = {
            'resultado': resultado,
            'apuestas_ganadoras': apuestas_ganadoras,
            'total_premios': total_premios,
        }
        return render(request, 'empresario/verificar_premios.html', context)
    
    def post(self, request, pk):
        if request.user.rol not in ['admin_general', 'empresario']:
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        resultado = get_object_or_404(ResultadoLoteria, pk=pk)
        
        if request.user.rol == 'empresario' and resultado.loteria.empresario != request.user:
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        # Marcar apuestas ganadoras como pagadas
        apuestas_ganadoras = Apuesta.objects.filter(
            loteria=resultado.loteria,
            fecha_hora__date=resultado.fecha,
            numero__endswith=resultado.numero_ganador[-resultado.cifras:],
            estado='activa',
            liquidacion_pagada=True
        )
        
        count = apuestas_ganadoras.update(estado='pagada')
        
        return JsonResponse({
            'success': True,
            'apuestas_pagadas': count
        })


@method_decorator(login_required, name='dispatch')
class ExportarVentasExcelView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        fecha_inicio = request.GET.get('fecha_inicio', (timezone.now() - timedelta(days=7)).strftime('%Y-%m-%d'))
        fecha_fin = request.GET.get('fecha_fin', timezone.now().strftime('%Y-%m-%d'))
        
        apuestas = Apuesta.objects.filter(
            empresario=request.user,
            fecha_hora__date__gte=fecha_inicio,
            fecha_hora__date__lte=fecha_fin
        ).select_related('chancero', 'loteria').order_by('-fecha_hora')
        
        # Crear archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ventas"
        
        # Encabezados
        headers = ['Fecha', 'Vendedor', 'Lotería', 'Número', 'Cifras', 'Monto', 'Premio', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        
        # Datos
        for row, apuesta in enumerate(apuestas, 2):
            ws.cell(row=row, column=1).value = apuesta.fecha_hora.strftime('%Y-%m-%d %H:%M')
            ws.cell(row=row, column=2).value = f"{apuesta.chancero.nombres} {apuesta.chancero.apellidos}"
            ws.cell(row=row, column=3).value = apuesta.loteria.nombre
            ws.cell(row=row, column=4).value = apuesta.numero
            ws.cell(row=row, column=5).value = apuesta.cifras
            ws.cell(row=row, column=6).value = float(apuesta.monto_apostado)
            ws.cell(row=row, column=7).value = float(apuesta.premio_potencial)
            ws.cell(row=row, column=8).value = apuesta.get_estado_display()
        
        # Ajustar ancho de columnas
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Crear respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=ventas_{fecha_inicio}_a_{fecha_fin}.xlsx'
        
        wb.save(response)
        return response


@method_decorator(login_required, name='dispatch')
class ExportarAcumuladosExcelView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        loteria_id = request.GET.get('loteria')
        
        acumulados = TopeNumero.objects.filter(
            empresario=request.user
        ).select_related('loteria')
        
        if loteria_id:
            acumulados = acumulados.filter(loteria_id=loteria_id)
        
        # Crear archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Acumulados"
        
        # Encabezados
        headers = ['Lotería', 'Número', 'Acumulado Actual', 'Tope Máximo', 'Disponible']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        
        # Datos
        for row, acum in enumerate(acumulados, 2):
            ws.cell(row=row, column=1).value = acum.loteria.nombre
            ws.cell(row=row, column=2).value = acum.numero
            ws.cell(row=row, column=3).value = float(acum.acumulado_actual)
            ws.cell(row=row, column=4).value = float(acum.tope_maximo)
            ws.cell(row=row, column=5).value = float(acum.tope_maximo - acum.acumulado_actual)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=acumulados.xlsx'
        
        wb.save(response)
        return response


@method_decorator(login_required, name='dispatch')
class LiquidacionesListView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')

        liquidaciones = Liquidacion.objects.filter(
            empresario=request.user
        ).select_related('chancero').order_by('-fecha_solicitud')

        # Calcular totales
        from decimal import Decimal
        total_ventas = liquidaciones.aggregate(total=Sum('total_ventas'))['total'] or Decimal('0')
        total_comisiones = liquidaciones.aggregate(total=Sum('comision_valor'))['total'] or Decimal('0')
        total_retenido = liquidaciones.aggregate(total=Sum('valor_empresario'))['total'] or Decimal('0')

        context = {
            'liquidaciones': liquidaciones,
            'total_ventas': total_ventas,
            'total_comisiones': total_comisiones,
            'total_retenido': total_retenido,
        }
        return render(request, 'empresario/liquidaciones.html', context)


@method_decorator(login_required, name='dispatch')
class LiquidacionCreateView(View):
    def post(self, request):
        if request.user.rol != 'empresario':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        chancero_id = request.POST.get('chancero_id')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        
        chancero = get_object_or_404(Usuario, pk=chancero_id, empresario=request.user)
        
        # Obtener comisión del chancero
        try:
            comision = ComisionVendedor.objects.get(
                empresario=request.user,
                chancero=chancero
            )
            porcentaje = comision.porcentaje
        except ComisionVendedor.DoesNotExist:
            return JsonResponse({'error': 'El chancero no tiene comisión asignada. Asigna una comisión antes de liquidar.'}, status=400)
        
        # Calcular ventas del período
        ventas = Apuesta.objects.filter(
            chancero=chancero,
            fecha_hora__date__gte=fecha_inicio,
            fecha_hora__date__lte=fecha_fin
        ).aggregate(total=Sum('monto_apostado'))['total'] or 0
        
        comision_valor = ventas * (porcentaje / 100)
        
        liquidacion = Liquidacion(
            empresario=request.user,
            chancero=chancero,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            total_ventas=ventas,
            comision_porcentaje=porcentaje,
            comision_valor=comision_valor,
            estado='solicitada'
        )
        liquidacion.save()
        
        return JsonResponse({
            'success': True,
            'id': liquidacion.id,
            'total_ventas': float(ventas),
            'comision_valor': float(comision_valor)
        })


@method_decorator(login_required, name='dispatch')
class LiquidacionPagarView(View):
    def post(self, request, pk):
        if request.user.rol != 'empresario':
            return JsonResponse({'error': 'No autorizado'}, status=403)

        liquidacion = get_object_or_404(Liquidacion, pk=pk, empresario=request.user)
        liquidacion.estado = 'pagada'
        liquidacion.fecha_pago = timezone.now()
        liquidacion.save()

        # Marcar las apuestas del período como liquidadas y pagadas
        # Esto autoriza el pago de premios para esas apuestas
        apuestas_periodo = Apuesta.objects.filter(
            empresario=request.user,
            chancero=liquidacion.chancero,
            fecha_hora__date__gte=liquidacion.fecha_inicio,
            fecha_hora__date__lte=liquidacion.fecha_fin,
            liquidacion_pagada=False
        )
        apuestas_periodo.update(
            liquidacion=liquidacion,
            liquidacion_pagada=True
        )

        # Generar factura de comisión global para el admin
        try:
            from decimal import Decimal
            comision_global = ComisionGlobal.objects.get(empresario=request.user, activo=True)
            valor_comision_admin = float(liquidacion.comision_valor) * (float(comision_global.porcentaje) / 100)

            FacturaComision.objects.create(
                empresario=request.user,
                fecha_inicio=liquidacion.fecha_inicio,
                fecha_fin=liquidacion.fecha_fin,
                total_ventas=liquidacion.total_ventas,
                porcentaje_comision=comision_global.porcentaje,
                valor_comision=Decimal(str(valor_comision_admin)),
                estado='pendiente'
            )
        except:
            pass

        from django.contrib import messages
        messages.success(request, f'Liquidación marcada como pagada. Comisión: ${liquidacion.comision_valor:,.0f}')
        return redirect('liquidaciones_list')


# ========== VISTAS PARA CHANCERO ==========

@method_decorator(login_required, name='dispatch')
class ApuestaCreateView(View):
    def get(self, request):
        if request.user.rol != 'chancero':
            return redirect('dashboard')

        # Verificar si el chancero está activo
        if not request.user.activo:
            return render(request, 'chancero/inactivo.html')

        # Loterías disponibles hoy
        ahora_dt = timezone.localtime(timezone.now())
        hoy = ahora_dt.strftime('%A')
        dias_map = {
            'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
            'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado',
            'Sunday': 'Domingo'
        }
        dia_hoy = dias_map[hoy]

        # Filtrar loterías activas del empresario
        loterias = Loteria.objects.filter(
            empresario=request.user.empresario,
            activa=True
        )

        loterias_filtradas = []
        hora_actual = ahora_dt.time()
        for loteria in loterias:
            dias = loteria.dias_habilitados or []
            if isinstance(dias, str):
                dias = [dias]
            dia_habilitado = dia_hoy in dias
            horario_valido = loteria.hora_apertura <= hora_actual <= loteria.hora_cierre
            if dia_habilitado and horario_valido:
                loterias_filtradas.append(loteria)

        context = {
            'loterias': loterias_filtradas,
        }
        return render(request, 'chancero/apuesta_form.html', context)
    
    def post(self, request):
        if request.user.rol != 'chancero':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        data = json.loads(request.body)
        loteria_id = data.get('loteria_id')
        numero = data.get('numero')
        monto = float(data.get('monto'))
        
        loteria = get_object_or_404(Loteria, pk=loteria_id, empresario=request.user.empresario)
        cifras = len(numero)
        
        # Validar rango de cifras configurado por empresario
        empresario = request.user.empresario
        cifras_min = empresario.cifras_minimas or 2
        cifras_max = empresario.cifras_maximas or 6
        
        if cifras < cifras_min or cifras > cifras_max:
            return JsonResponse({
                'error': f'El número de cifras debe estar entre {cifras_min} y {cifras_max}. Actual: {cifras}'
            }, status=400)
        
        # Validar hora
        ahora = timezone.localtime(timezone.now()).time()
        if not (loteria.hora_apertura <= ahora <= loteria.hora_cierre):
            return JsonResponse({
                'error': f'Fuera de horario. Horario: {loteria.hora_apertura} - {loteria.hora_cierre}'
            }, status=400)

        # Obtener premio por peso
        try:
            plan = PlanPremio.objects.get(
                empresario=request.user.empresario,
                cifras=cifras,
                es_combinado=False
            )
            premio = monto * float(plan.premio_por_peso)
        except PlanPremio.DoesNotExist:
            return JsonResponse({'error': f'No hay plan de premios configurado para {cifras} cifras'}, status=400)

        # Validar tope por LOTERÍA (NO por número)
        try:
            tope = TopeNumero.objects.get(
                empresario=request.user.empresario,
                loteria=loteria
            )
            # Usar el acumulado_actual del modelo para consistencia
            acumulado = float(tope.acumulado_actual or 0)

            if acumulado + premio > float(tope.tope_maximo):
                maximo_permitido = float(tope.tope_maximo) - acumulado
                return JsonResponse({
                    'error': f'El tope de premios para esta lotería es ${float(tope.tope_maximo):,.0f}. '
                            f'Actualmente acumulado hoy: ${acumulado:,.0f}. '
                            f'Máximo permitido para esta apuesta: ${maximo_permitido:,.0f}'
                }, status=400)
        except TopeNumero.DoesNotExist:
            pass  # No hay tope configurado
        
        # Crear apuesta
        apuesta = Apuesta(
            empresario=request.user.empresario,
            chancero=request.user,
            loteria=loteria,
            numero=numero,
            cifras=cifras,
            monto_apostado=monto,
            premio_potencial=premio,
            estado='activa'
        )
        apuesta.save()
        
        # Actualizar acumulado del tope (si existe)
        try:
            tope = TopeNumero.objects.get(
                empresario=request.user.empresario,
                loteria=loteria
            )
            tope.acumulado_actual = models.F('acumulado_actual') + premio
            tope.save()
        except TopeNumero.DoesNotExist:
            pass
        
        return JsonResponse({
            'success': True,
            'id': apuesta.id,
            'premio': float(premio),
            'mensaje': 'Apuesta registrada exitosamente'
        })


    
    def post(self, request):
        if request.user.rol != 'chancero':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        data = json.loads(request.body)
        loteria_id = data.get('loteria_id')
        numero = data.get('numero')
        monto = float(data.get('monto'))
        
        loteria = get_object_or_404(Loteria, pk=loteria_id, empresario=request.user.empresario)
        cifras = len(numero)
        
        # Validar hora
        ahora = timezone.now().time()
        if not (loteria.hora_apertura <= ahora <= loteria.hora_cierre):
            return JsonResponse({
                'error': f'Fuera de horario. Horario: {loteria.hora_apertura} - {loteria.hora_cierre}'
            }, status=400)
        
        # Obtener multiplicador
        try:
            plan = PlanPremio.objects.get(
                empresario=request.user.empresario,
                loteria=loteria,
                cifras=cifras
            )
            premio = monto * float(plan.multiplicador)
        except:
            return JsonResponse({'error': 'No hay plan de premios para esta lotería'}, status=400)
        
        # Validar tope
        try:
            tope = TopeNumero.objects.get(
                empresario=request.user.empresario,
                loteria=loteria,
                numero=numero
            )
            if float(tope.acumulado_actual) + monto > float(tope.tope_maximo):
                maximo_permitido = float(tope.tope_maximo) - float(tope.acumulado_actual)
                return JsonResponse({
                    'error': f'Tope excedido. Máximo permitido: ${maximo_permitido:,.0f}'
                }, status=400)
        except TopeNumero.DoesNotExist:
            pass
        
        # Crear apuesta
        apuesta = Apuesta(
            empresario=request.user.empresario,
            chancero=request.user,
            loteria=loteria,
            numero=numero,
            cifras=cifras,
            monto_apostado=monto,
            premio_potencial=premio,
            estado='activa'
        )
        apuesta.save()
        
        # Actualizar acumulado
        TopeNumero.objects.update_or_create(
            empresario=request.user.empresario,
            loteria=loteria,
            numero=numero,
            defaults={
                'tope_maximo': tope.tope_maximo if 'tope' in locals() else 999999999,
                'acumulado_actual': models.F('acumulado_actual') + monto
            }
        )
        
        return JsonResponse({
            'success': True,
            'id': apuesta.id,
            'premio': float(premio),
            'mensaje': 'Apuesta registrada exitosamente'
        })


@method_decorator(login_required, name='dispatch')
class GenerarReporteApuestasView(View):
    def post(self, request):
        if request.user.rol != 'chancero':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        data = json.loads(request.body)
        apuestas_ids = data.get('apuestas_ids', [])
        telefono = data.get('telefono', '')
        
        if not apuestas_ids:
            return JsonResponse({'error': 'No se proporcionaron apuestas'}, status=400)
        
        apuestas = Apuesta.objects.filter(
            id__in=apuestas_ids,
            chancero=request.user
        ).select_related('loteria')
        
        if not apuestas:
            return JsonResponse({'error': 'No se encontraron las apuestas'}, status=400)
        
        # Generar texto del reporte
        fecha_hora = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        
        reporte_lines = [
            f"🎰 CHANCE PRO - REPORTE DE APUESTAS",
            f"📅 Fecha: {fecha_hora}",
            f"👤 Chancero: {request.user.nombres} {request.user.apellidos}",
            f"📞 Tel: {request.user.telefono}",
            "",
            "📋 APUESTAS:",
        ]
        
        total_monto = 0
        total_premio = 0
        
        for apuesta in apuestas:
            reporte_lines.append(
                f"• {apuesta.loteria.nombre} | N°: {apuesta.numero} | "
                f"${apuesta.monto_apostado:,.0f} | Premio: ${apuesta.premio_potencial:,.0f}"
            )
            total_monto += float(apuesta.monto_apostado)
            total_premio += float(apuesta.premio_potencial)
        
        reporte_lines.extend([
            "",
            f"💰 Total apostado: ${total_monto:,.0f}",
            f"🏆 Premio potencial total: ${total_premio:,.0f}",
            "",
            "¡Buena suerte! 🍀"
        ])
        
        reporte_text = "\\n".join(reporte_lines)
        
        # Guardar mensaje en la base de datos
        if telefono:
            Mensaje.objects.create(
                empresario=request.user.empresario,
                chancero=request.user,
                telefono_destino=telefono,
                contenido=reporte_text,
                tipo='whatsapp'
            )
        
        return JsonResponse({
            'success': True,
            'reporte': reporte_text,
            'total_monto': total_monto,
            'total_premio': total_premio
        })


@method_decorator(login_required, name='dispatch')
class VerificarDisponibilidadView(View):
    def get(self, request):
        if request.user.rol != 'chancero':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        loteria_id = request.GET.get('loteria_id')
        numero = request.GET.get('numero')
        monto = float(request.GET.get('monto', 0))
        
        try:
            tope = TopeNumero.objects.get(
                empresario=request.user.empresario,
                loteria_id=loteria_id,
                numero=numero
            )
            disponible = float(tope.tope_maximo) - float(tope.acumulado_actual)
            return JsonResponse({
                'disponible': disponible,
                'tope': float(tope.tope_maximo),
                'acumulado': float(tope.acumulado_actual),
                'permitido': monto <= disponible
            })
        except TopeNumero.DoesNotExist:
            return JsonResponse({
                'disponible': 999999999,
                'tope': 999999999,
                'acumulado': 0,
                'permitido': True
            })


@method_decorator(login_required, name='dispatch')
class MisApuestasListView(View):
    def get(self, request):
        if request.user.rol != 'chancero':
            return redirect('dashboard')
        
        fecha = request.GET.get('fecha', timezone.now().strftime('%Y-%m-%d'))
        
        apuestas = Apuesta.objects.filter(
            chancero=request.user,
            fecha_hora__date=fecha
        ).select_related('loteria').order_by('-fecha_hora')
        
        # Cálculos de recaudo
        total_dia = apuestas.aggregate(total=Sum('monto_apostado'))['total'] or 0
        apuestas_count = apuestas.count()
        
        # Premios ganados (pagados)
        premios_ganados = apuestas.filter(estado='pagada').aggregate(
            total=Sum('premio_potencial')
        )['total'] or 0
        
        # Obtener porcentaje de comisión del chancero
        try:
            comision = ComisionVendedor.objects.get(chancero=request.user)
            porcentaje_comision = comision.porcentaje
        except:
            porcentaje_comision = 0
        
        # Calcular ganancia del chancero
        ganancia_estimada = total_dia * (porcentaje_comision / 100)
        
        # Apuestas por lotería
        ventas_por_loteria = apuestas.values('loteria__nombre').annotate(
            total=Sum('monto_apostado'),
            count=Count('id')
        ).order_by('-total')
        
        context = {
            'apuestas': apuestas,
            'fecha': fecha,
            'total_dia': total_dia,
            'apuestas_count': apuestas_count,
            'premios_ganados': premios_ganados,
            'porcentaje_comision': porcentaje_comision,
            'ganancia_estimada': ganancia_estimada,
            'ventas_por_loteria': ventas_por_loteria,
        }
        return render(request, 'chancero/mis_apuestas.html', context)


@method_decorator(login_required, name='dispatch')
class EnviarMensajeView(View):
    def post(self, request):
        if request.user.rol != 'chancero':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        data = json.loads(request.body)
        apuesta_id = data.get('apuesta_id')
        telefono = data.get('telefono')
        tipo = data.get('tipo', 'whatsapp')
        
        apuesta = get_object_or_404(Apuesta, pk=apuesta_id, chancero=request.user)
        
        # Crear mensaje (simulado)
        contenido = f"ChancePro - Apuesta registrada\n"
        contenido += f"Lotería: {apuesta.loteria.nombre}\n"
        contenido += f"Número: {apuesta.numero}\n"
        contenido += f"Monto: ${apuesta.monto_apostado:,.0f}\n"
        contenido += f"Premio: ${apuesta.premio_potencial:,.0f}\n"
        contenido += f"Fecha: {apuesta.fecha_hora.strftime('%Y-%m-%d %H:%M')}"
        
        mensaje = Mensaje(
            empresario=request.user.empresario,
            chancero=request.user,
            apuesta=apuesta,
            telefono_destino=telefono,
            contenido=contenido,
            tipo=tipo
        )
        mensaje.save()
        
        # Simular envío (en producción aquí iría la integración real)
        return JsonResponse({
            'success': True,
            'mensaje': 'Mensaje enviado correctamente',
            'contenido': contenido
        })


@method_decorator(login_required, name='dispatch')
class MiLiquidacionView(View):
    def get(self, request):
        if request.user.rol != 'chancero':
            return redirect('dashboard')

        liquidaciones = Liquidacion.objects.filter(
            chancero=request.user
        ).order_by('-fecha_solicitud')

        # Calcular totales del chancero
        from decimal import Decimal
        total_ventas = liquidaciones.aggregate(total=Sum('total_ventas'))['total'] or Decimal('0')
        total_comisiones = liquidaciones.aggregate(total=Sum('comision_valor'))['total'] or Decimal('0')

        context = {
            'liquidaciones': liquidaciones,
            'total_ventas': total_ventas,
            'total_comisiones': total_comisiones,
        }
        return render(request, 'chancero/mis_liquidaciones.html', context)


@method_decorator(login_required, name='dispatch')
class LoteriaDetailView(View):
    def get(self, request, pk):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        loteria = get_object_or_404(Loteria, pk=pk, empresario=request.user)
        
        context = {
            'loteria': loteria,
        }
        return render(request, 'empresario/loteria_detail.html', context)


@method_decorator(login_required, name='dispatch')
class LiquidacionSolicitarView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')

        hoy = timezone.now().date()
        fecha_inicio = request.GET.get('fecha_inicio', hoy.strftime('%Y-%m-%d'))
        fecha_fin = request.GET.get('fecha_fin', hoy.strftime('%Y-%m-%d'))
        chanceros = Usuario.objects.filter(rol='chancero', empresario=request.user).order_by('nombres', 'apellidos')
        chanceros_data = []

        from decimal import Decimal

        for chancero in chanceros:
            ventas = Apuesta.objects.filter(
                empresario=request.user,
                chancero=chancero,
                fecha_hora__date__gte=fecha_inicio,
                fecha_hora__date__lte=fecha_fin,
                liquidacion_pagada=False
            ).aggregate(total=Sum('monto_apostado'))['total'] or Decimal('0')
            apuestas_count = Apuesta.objects.filter(
                empresario=request.user,
                chancero=chancero,
                fecha_hora__date__gte=fecha_inicio,
                fecha_hora__date__lte=fecha_fin,
                liquidacion_pagada=False
            ).count()

            try:
                comision = ComisionVendedor.objects.get(empresario=request.user, chancero=chancero)
                porcentaje = float(comision.porcentaje)
            except ComisionVendedor.DoesNotExist:
                porcentaje = 0.0

            comision_valor = float(ventas) * (porcentaje / 100) if porcentaje > 0 else Decimal('0')
            valor_empresario = float(ventas) - float(comision_valor)

            chanceros_data.append({
                'chancero': chancero,
                'ventas': ventas,
                'apuestas_count': apuestas_count,
                'porcentaje': porcentaje,
                'comision_valor': comision_valor,
                'valor_empresario': valor_empresario,
            })

        context = {
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'chanceros_data': chanceros_data,
        }
        return render(request, 'empresario/liquidacion_solicitar.html', context)
    
    def post(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')

        chancero_id = request.POST.get('chancero_id')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        chancero = get_object_or_404(Usuario, pk=chancero_id, empresario=request.user, rol='chancero')

        from decimal import Decimal

        try:
            comision = ComisionVendedor.objects.get(empresario=request.user, chancero=chancero)
            porcentaje = Decimal(str(comision.porcentaje))
        except ComisionVendedor.DoesNotExist:
            porcentaje = Decimal('0')

        ventas = Apuesta.objects.filter(
            empresario=request.user,
            chancero=chancero,
            fecha_hora__date__gte=fecha_inicio,
            fecha_hora__date__lte=fecha_fin,
            liquidacion_pagada=False
        ).aggregate(total=Sum('monto_apostado'))['total'] or Decimal('0')

        comision_valor = ventas * (porcentaje / Decimal('100')) if porcentaje > 0 else Decimal('0')
        valor_empresario = ventas - comision_valor

        liquidacion = Liquidacion.objects.create(
            empresario=request.user,
            chancero=chancero,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            total_ventas=ventas,
            comision_porcentaje=porcentaje,
            comision_valor=comision_valor,
            valor_empresario=valor_empresario,
            estado='solicitada',
        )

        Apuesta.objects.filter(
            empresario=request.user,
            chancero=chancero,
            fecha_hora__date__gte=fecha_inicio,
            fecha_hora__date__lte=fecha_fin,
            liquidacion_pagada=False
        ).update(liquidacion=liquidacion)

        from django.contrib import messages
        messages.success(request, f'Chancero {chancero.nombres} {chancero.apellidos} liquidado exitosamente. Comisión: ${comision_valor:,.0f}')

        return redirect('liquidaciones_list')


@method_decorator(login_required, name='dispatch')
class EmpresarioConfigView(View):
    def get(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        context = {
            'empresario': request.user,
        }
        return render(request, 'empresario/config.html', context)
    
    def post(self, request):
        if request.user.rol != 'empresario':
            return redirect('dashboard')
        
        empresario = request.user
        empresario.autorizar_2c_directo = request.POST.get('autorizar_2c_directo') == 'on'
        empresario.autorizar_2c_combinado = request.POST.get('autorizar_2c_combinado') == 'on'
        empresario.autorizar_3c_directo = request.POST.get('autorizar_3c_directo') == 'on'
        empresario.autorizar_3c_combinado = request.POST.get('autorizar_3c_combinado') == 'on'
        empresario.autorizar_4c_directo = request.POST.get('autorizar_4c_directo') == 'on'
        empresario.autorizar_4c_combinado = request.POST.get('autorizar_4c_combinado') == 'on'
        empresario.autorizar_5c_directo = request.POST.get('autorizar_5c_directo') == 'on'
        empresario.autorizar_5c_combinado = request.POST.get('autorizar_5c_combinado') == 'on'
        empresario.autorizar_6c_directo = request.POST.get('autorizar_6c_directo') == 'on'
        empresario.autorizar_6c_combinado = request.POST.get('autorizar_6c_combinado') == 'on'
        empresario.save()
        
        return render(request, 'empresario/config.html', {
            'empresario': empresario,
            'success': 'Configuración actualizada exitosamente'
        })


# ========== VISTAS PARA AGENDA DE CLIENTES ==========

@method_decorator(login_required, name='dispatch')
class AgendaClientesView(View):
    def get(self, request):
        if request.user.rol != 'chancero':
            return redirect('dashboard')
        
        clientes = Cliente.objects.filter(
            chancero=request.user,
            activo=True
        ).order_by('nombre')
        
        context = {
            'clientes': clientes,
        }
        return render(request, 'chancero/agenda_clientes.html', context)


@method_decorator(login_required, name='dispatch')
class ClienteCreateView(View):
    def post(self, request):
        if request.user.rol != 'chancero':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        nombre = request.POST.get('nombre')
        telefono = request.POST.get('telefono')
        direccion = request.POST.get('direccion', '')
        notas = request.POST.get('notas', '')
        
        if not nombre or not telefono:
            return JsonResponse({'error': 'Nombre y teléfono son obligatorios'}, status=400)
        
        # Verificar si ya existe el teléfono
        if Cliente.objects.filter(telefono=telefono).exists():
            return JsonResponse({'error': 'El teléfono ya está registrado'}, status=400)
        
        cliente = Cliente.objects.create(
            chancero=request.user,
            nombre=nombre,
            telefono=telefono,
            direccion=direccion,
            notas=notas
        )
        
        return JsonResponse({'success': True, 'id': cliente.id})


@method_decorator(login_required, name='dispatch')
class ClienteUpdateView(View):
    def post(self, request, pk):
        if request.user.rol != 'chancero':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        cliente = get_object_or_404(Cliente, pk=pk, chancero=request.user)
        
        cliente.nombre = request.POST.get('nombre', cliente.nombre)
        cliente.telefono = request.POST.get('telefono', cliente.telefono)
        cliente.direccion = request.POST.get('direccion', cliente.direccion)
        cliente.notas = request.POST.get('notas', cliente.notas)
        cliente.save()
        
        return JsonResponse({'success': True})


@method_decorator(login_required, name='dispatch')
class ClienteDeleteView(View):
    def post(self, request, pk):
        if request.user.rol != 'chancero':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        cliente = get_object_or_404(Cliente, pk=pk, chancero=request.user)
        cliente.activo = False
        cliente.save()
        
        return JsonResponse({'success': True})


@method_decorator(login_required, name='dispatch')
class ClientesAutocompleteView(View):
    def get(self, request):
        if request.user.rol != 'chancero':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        query = request.GET.get('q', '')
        
        clientes = Cliente.objects.filter(
            chancero=request.user,
            activo=True
        ).filter(
            Q(nombre__icontains=query) | Q(telefono__icontains=query)
        )[:10]
        
        results = []
        for cliente in clientes:
            results.append({
                'id': cliente.id,
                'nombre': cliente.nombre,
                'telefono': cliente.telefono,
                'label': f"{cliente.nombre} - {cliente.telefono}"
            })
        
        return JsonResponse({'results': results})


@method_decorator(login_required, name='dispatch')
class TopesLoteriasAPIView(View):
    def get(self, request):
        if request.user.rol != 'chancero':
            return JsonResponse({'error': 'No autorizado'}, status=403)
        
        topes = TopeNumero.objects.filter(
            empresario=request.user.empresario
        ).select_related('loteria')
        
        topes_data = []
        for tope in topes:
            topes_data.append({
                'loteria_id': tope.loteria.id,
                'loteria_nombre': tope.loteria.nombre,
                'tope_maximo': float(tope.tope_maximo),
                'acumulado_actual': float(tope.acumulado_actual),
            })
        
        return JsonResponse(topes_data, safe=False)


@method_decorator(login_required, name='dispatch')
class HistorialClienteView(View):
    def get(self, request, pk):
        if request.user.rol != 'chancero':
            return redirect('dashboard')
        
        cliente = get_object_or_404(Cliente, pk=pk, chancero=request.user)
        
        historial = HistorialCliente.objects.filter(
            cliente=cliente
        ).select_related('apuesta__loteria').order_by('-fecha_registro')
        
        # Calcular estadísticas
        total_apuestas = historial.count()
        total_apostado = sum(h.apuesta.monto_apostado for h in historial)
        total_premios = sum(h.apuesta.premio_potencial for h in historial)
        
        # Números más jugados
        from collections import Counter
        numeros_jugados = [h.apuesta.numero for h in historial]
        numeros_frecuentes = Counter(numeros_jugados).most_common(5)
        
        context = {
            'cliente': cliente,
            'historial': historial,
            'total_apuestas': total_apuestas,
            'total_apostado': total_apostado,
            'total_premios': total_premios,
            'numeros_frecuentes': numeros_frecuentes,
        }
        return render(request, 'chancero/historial_cliente.html', context)


@method_decorator(login_required, name='dispatch')
class EstadisticasPersonalesView(View):
    def get(self, request):
        if request.user.rol != 'chancero':
            return redirect('dashboard')
        
        # Obtener parámetros de fecha
        fecha_inicio = request.GET.get('fecha_inicio', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
        fecha_fin = request.GET.get('fecha_fin', timezone.now().strftime('%Y-%m-%d'))
        
        # Apuestas del período
        apuestas = Apuesta.objects.filter(
            chancero=request.user,
            fecha_hora__date__gte=fecha_inicio,
            fecha_hora__date__lte=fecha_fin
        ).select_related('loteria')
        
        # Estadísticas generales
        total_apuestas = apuestas.count()
        total_ventas = apuestas.aggregate(total=Sum('monto_apostado'))['total'] or 0
        total_premios_potenciales = apuestas.aggregate(total=Sum('premio_potencial'))['total'] or 0
        
        # Obtener comisión
        try:
            comision = ComisionVendedor.objects.get(chancero=request.user)
            porcentaje_comision = comision.porcentaje
            comision_estimada = total_ventas * (porcentaje_comision / 100)
        except ComisionVendedor.DoesNotExist:
            porcentaje_comision = 0
            comision_estimada = 0
        
        # Números más jugados
        from collections import Counter
        numeros_jugados = [apuesta.numero for apuesta in apuestas]
        numeros_frecuentes = Counter(numeros_jugados).most_common(10)
        
        # Loterías favoritas
        loterias_jugadas = [apuesta.loteria.nombre for apuesta in apuestas]
        loterias_frecuentes = Counter(loterias_jugadas).most_common(5)
        
        # Ventas por día de la semana
        ventas_por_dia = {}
        dias_map = {0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves', 4: 'Viernes', 5: 'Sábado', 6: 'Domingo'}
        
        for apuesta in apuestas:
            dia = apuesta.fecha_hora.weekday()
            dia_nombre = dias_map[dia]
            if dia_nombre not in ventas_por_dia:
                ventas_por_dia[dia_nombre] = 0
            ventas_por_dia[dia_nombre] += float(apuesta.monto_apostado)
        
        # Ventas por fecha (últimos 7 días)
        ventas_por_fecha = {}
        for i in range(7):
            fecha = (timezone.now() - timedelta(days=i)).date()
            ventas_dia = apuestas.filter(fecha_hora__date=fecha).aggregate(total=Sum('monto_apostado'))['total'] or 0
            ventas_por_fecha[fecha.strftime('%Y-%m-%d')] = float(ventas_dia)
        
        # Apuestas por cifras
        apuestas_por_cifras = {}
        for apuesta in apuestas:
            cifras = apuesta.cifras
            if cifras not in apuestas_por_cifras:
                apuestas_por_cifras[cifras] = 0
            apuestas_por_cifras[cifras] += 1
        
        context = {
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'total_apuestas': total_apuestas,
            'total_ventas': total_ventas,
            'total_premios_potenciales': total_premios_potenciales,
            'porcentaje_comision': porcentaje_comision,
            'comision_estimada': comision_estimada,
            'numeros_frecuentes': numeros_frecuentes,
            'loterias_frecuentes': loterias_frecuentes,
            'ventas_por_dia': ventas_por_dia,
            'ventas_por_fecha': ventas_por_fecha,
            'apuestas_por_cifras': apuestas_por_cifras,
        }
        return render(request, 'chancero/estadisticas_personales.html', context)


@method_decorator(login_required, name='dispatch')
class GananciasDiaView(View):
    def get(self, request):
        if request.user.rol != 'chancero':
            return redirect('dashboard')
        
        fecha = request.GET.get('fecha', timezone.now().strftime('%Y-%m-%d'))
        
        # Apuestas del día
        apuestas_hoy = Apuesta.objects.filter(
            chancero=request.user,
            fecha_hora__date=fecha
        ).select_related('loteria')
        
        # Calcular ventas del día
        ventas_hoy = apuestas_hoy.aggregate(total=Sum('monto_apostado'))['total'] or 0
        total_apuestas = apuestas_hoy.count()
        
        # Obtener comisión
        try:
            comision = ComisionVendedor.objects.get(chancero=request.user)
            porcentaje_comision = comision.porcentaje
            ganancia_hoy = ventas_hoy * (porcentaje_comision / 100)
        except ComisionVendedor.DoesNotExist:
            porcentaje_comision = 0
            ganancia_hoy = 0
        
        # Ventas por lotería
        ventas_por_loteria = {}
        for apuesta in apuestas_hoy:
            loteria_nombre = apuesta.loteria.nombre
            if loteria_nombre not in ventas_por_loteria:
                ventas_por_loteria[loteria_nombre] = 0
            ventas_por_loteria[loteria_nombre] += float(apuesta.monto_apostado)
        
        # Ventas por hora
        ventas_por_hora = {}
        for apuesta in apuestas_hoy:
            hora = apuesta.fecha_hora.hour
            if hora not in ventas_por_hora:
                ventas_por_hora[hora] = 0
            ventas_por_hora[hora] += float(apuesta.monto_apostado)
        
        # Últimas apuestas del día
        ultimas_apuestas = apuestas_hoy.order_by('-fecha_hora')[:10]
        
        context = {
            'fecha': fecha,
            'ventas_hoy': ventas_hoy,
            'total_apuestas': total_apuestas,
            'porcentaje_comision': porcentaje_comision,
            'ganancia_hoy': ganancia_hoy,
            'ventas_por_loteria': ventas_por_loteria,
            'ventas_por_hora': ventas_por_hora,
            'ultimas_apuestas': ultimas_apuestas,
        }
        return render(request, 'chancero/ganancias_dia.html', context)


@method_decorator(login_required, name='dispatch')
class InformacionChanceroView(View):
    def get(self, request):
        if request.user.rol != 'chancero':
            return redirect('dashboard')
        
        hoy = timezone.now().date()
        
        # Ventas de hoy
        ventas_hoy = Apuesta.objects.filter(
            chancero=request.user,
            fecha_hora__date=hoy
        ).aggregate(total=Sum('monto_apostado'))['total'] or 0
        
        apuestas_hoy = Apuesta.objects.filter(
            chancero=request.user,
            fecha_hora__date=hoy
        ).count()
        
        # Obtener comisión
        try:
            comision = ComisionVendedor.objects.get(chancero=request.user)
            porcentaje_comision = comision.porcentaje
            ganancia_hoy = ventas_hoy * (porcentaje_comision / 100)
        except ComisionVendedor.DoesNotExist:
            porcentaje_comision = 0
            ganancia_hoy = 0
        
        # Ventas del mes
        ventas_mes = Apuesta.objects.filter(
            chancero=request.user,
            fecha_hora__month=hoy.month,
            fecha_hora__year=hoy.year
        ).aggregate(total=Sum('monto_apostado'))['total'] or 0
        
        ganancia_mes = ventas_mes * (porcentaje_comision / 100)
        
        # Últimas apuestas
        ultimas_apuestas = Apuesta.objects.filter(
            chancero=request.user
        ).select_related('loteria').order_by('-fecha_hora')[:10]
        
        context = {
            'ventas_hoy': ventas_hoy,
            'apuestas_hoy': apuestas_hoy,
            'porcentaje_comision': porcentaje_comision,
            'ganancia_hoy': ganancia_hoy,
            'ventas_mes': ventas_mes,
            'ganancia_mes': ganancia_mes,
            'ultimas_apuestas': ultimas_apuestas,
        }
        return render(request, 'chancero/informacion_chancero.html', context)


@method_decorator(login_required, name='dispatch')
class GenerarTicketView(View):
    def get(self, request, apuesta_id):
        if request.user.rol != 'chancero':
            return redirect('dashboard')
        
        apuesta = get_object_or_404(Apuesta, pk=apuesta_id, chancero=request.user)
        
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from io import BytesIO
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title = Paragraph("TICKET DE APUESTA", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))
        
        # Información de la apuesta
        data = [
            ['Fecha:', apuesta.fecha_hora.strftime('%Y-%m-%d %H:%M')],
            ['Lotería:', apuesta.loteria.nombre],
            ['Número:', apuesta.numero],
            ['Cifras:', str(apuesta.cifras)],
            ['Monto Apostado:', f'${float(apuesta.monto_apostado):.2f}'],
            ['Premio Potencial:', f'${float(apuesta.premio_potencial):.2f}'],
            ['Estado:', apuesta.estado.upper()],
            ['Chancero:', f'{request.user.nombres} {request.user.apellidos}'],
            ['Empresario:', f'{apuesta.empresario.nombres} {apuesta.empresario.apellidos}'],
        ]
        
        table = Table(data, colWidths=[2 * inch, 3 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.5 * inch))
        
        # Advertencia
        warning = Paragraph("Este ticket es prueba de su apuesta. Guárdelo para reclamar premios.", styles['Normal'])
        elements.append(warning)
        
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=ticket_{apuesta_id}.pdf'
        return response


@method_decorator(login_required, name='dispatch')
class GenerarQRView(View):
    def get(self, request, apuesta_id):
        if request.user.rol != 'chancero':
            return redirect('dashboard')
        
        apuesta = get_object_or_404(Apuesta, pk=apuesta_id, chancero=request.user)
        
        import qrcode
        from io import BytesIO
        
        # Crear datos del QR
        qr_data = {
            'apuesta_id': apuesta.id,
            'fecha': apuesta.fecha_hora.strftime('%Y-%m-%d %H:%M'),
            'loteria': apuesta.loteria.nombre,
            'numero': apuesta.numero,
            'cifras': apuesta.cifras,
            'monto': str(float(apuesta.monto_apostado)),
            'premio': str(float(apuesta.premio_potencial)),
            'chancero': f'{request.user.nombres} {request.user.apellidos}',
            'empresario': f'{apuesta.empresario.nombres} {apuesta.empresario.apellidos}',
        }
        
        import json
        qr_data_str = json.dumps(qr_data)
        
        # Generar QR
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data_str)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        
        response = HttpResponse(buffer.read(), content_type='image/png')
        response['Content-Disposition'] = f'attachment; filename=qr_{apuesta_id}.png'
        return response
